import asyncio
import csv
import logging
import os
import time
import traceback
import aiofiles
from datetime import datetime, timedelta, date
from io import BytesIO, StringIO
from typing import Dict, List, Optional

import openpyxl
from aiogram import types
from aiogram.types import FSInputFile
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from bot_manager import bot_manager
from config import Config, beijing_tz
from database import db
from decorators import admin_required
from fault_tolerance import Watchdog
from keyboards import is_admin
from performance import track_performance, with_retry, message_deduplicate, rate_limit
from utils import MessageFormatter, notification_service, user_lock_manager
from handover_manager import handover_manager

logger = logging.getLogger("GroupCheckInBot")

# ========== 生成月度报告函数 =========
async def generate_monthly_report(chat_id: int, year: int = None, month: int = None):
    """生成月度报告"""
    if year is None or month is None:
        today = db.get_beijing_time()
        year = today.year
        month = today.month

    monthly_stats = await db.get_monthly_statistics(chat_id, year, month)
    work_stats = await db.get_monthly_work_statistics(chat_id, year, month)
    activity_ranking = await db.get_monthly_activity_ranking(chat_id, year, month)

    if not monthly_stats and not work_stats:
        return None

    chat_title = str(chat_id)
    try:
        if bot_manager.bot:
            chat_info = await bot_manager.bot.get_chat(chat_id)
            chat_title = chat_info.title or chat_title
    except Exception:
        pass

    report = (
        f"📊 <b>{year}年{month}月打卡统计报告</b>\n"
        f"🏢 群组：<code>{chat_title}</code>\n"
        f"📅 生成时间：<code>{db.get_beijing_time().strftime('%Y-%m-%d %H:%M:%S')}</code>\n"
        f"{MessageFormatter.create_dashed_line()}\n"
    )

    total_users = len(monthly_stats)
    total_activity_time = sum(
        stat.get("total_accumulated_time", 0) for stat in monthly_stats
    )
    total_activity_count = sum(
        stat.get("total_activity_count", 0) for stat in monthly_stats
    )
    total_fines = sum(stat.get("total_fines", 0) for stat in monthly_stats)

    total_work_days = sum(stat.get("work_days", 0) for stat in monthly_stats)
    total_work_hours = sum(stat.get("work_hours", 0) for stat in monthly_stats)

    report += (
        f"👥 <b>总体统计</b>\n"
        f"• 活跃用户：<code>{total_users}</code> 人\n"
        f"• 总活动时长：<code>{MessageFormatter.format_time(int(total_activity_time))}</code>\n"
        f"• 总活动次数：<code>{total_activity_count}</code> 次\n"
        f"• 总工作天数：<code>{total_work_days}</code> 天\n"
        f"• 总工作时长：<code>{MessageFormatter.format_time(int(total_work_hours))}</code>\n"
        f"• 总扣罚款金额：<code>{total_fines}</code> 泰铢\n\n"
    )

    total_work_start = sum(stat.get("work_start_count", 0) for stat in work_stats)
    total_work_end = sum(stat.get("work_end_count", 0) for stat in work_stats)
    total_work_fines = sum(
        stat.get("work_start_fines", 0) + stat.get("work_end_fines", 0)
        for stat in work_stats
    )

    if total_work_start > 0 or total_work_end > 0:
        report += (
            f"🕒 <b>上下班统计</b>\n"
            f"• 上班打卡：<code>{total_work_start}</code> 次\n"
            f"• 下班打卡：<code>{total_work_end}</code> 次\n"
            f"• 上下班罚款：<code>{total_work_fines}</code> 泰铢\n\n"
        )

    if monthly_stats:
        report += f"👤 <b>个人工作统计</b>\n"

        work_hours_ranking = sorted(
            [stat for stat in monthly_stats if stat.get("work_hours", 0) > 0],
            key=lambda x: x.get("work_hours", 0),
            reverse=True,
        )[:5]

        for i, stat in enumerate(work_hours_ranking, 1):
            work_hours_str = MessageFormatter.format_time(
                int(stat.get("work_hours", 0))
            )
            work_days = stat.get("work_days", 0)
            nickname = stat.get("nickname", f"用户{stat.get('user_id')}")
            report += (
                f"  <code>{i}.</code> {nickname} - {work_hours_str} ({work_days}天)\n"
            )
        report += "\n"

    report += f"🏆 <b>月度活动排行榜</b>\n"
    has_activity_data = False

    for activity, ranking in activity_ranking.items():
        if ranking:
            has_activity_data = True
            report += f"📈 <code>{activity}</code>：\n"
            for i, user in enumerate(ranking[:3], 1):
                time_str = MessageFormatter.format_time(int(user.get("total_time", 0)))
                count = user.get("total_count", 0)
                nickname = user.get("nickname", "未知用户")
                report += f"  <code>{i}.</code> {nickname} - {time_str} ({count}次)\n"
            report += "\n"

    if not has_activity_data:
        report += "暂无活动数据\n\n"

    report += f"📈 <b>月度总结</b>\n"

    if total_activity_count > 0:
        avg_activity_time = (
            total_activity_time / total_activity_count
            if total_activity_count > 0
            else 0
        )
        report += f"• 平均每次活动时长：<code>{MessageFormatter.format_time(int(avg_activity_time))}</code>\n"

    if total_work_days > 0:
        avg_work_hours_per_day = (
            total_work_hours / total_work_days if total_work_days > 0 else 0
        )
        report += f"• 平均每日工作时长：<code>{MessageFormatter.format_time(int(avg_work_hours_per_day))}</code>\n"

    if total_users > 0:
        avg_activity_per_user = (
            total_activity_count / total_users if total_users > 0 else 0
        )
        report += f"• 人均活动次数：<code>{avg_activity_per_user:.1f}</code> 次\n"

        avg_work_days_per_user = total_work_days / total_users if total_users > 0 else 0
        report += f"• 人均工作天数：<code>{avg_work_days_per_user:.1f}</code> 天\n"

    report += f"\n{MessageFormatter.create_dashed_line()}\n"
    report += f"💡 <i>注：本报告基于月度统计表生成，不受日常重置操作影响</i>"

    return report


async def ensure_monthly_data_completeness(stats: List[Dict]) -> List[Dict]:
    """确保月度统计数据的完整性"""
    if not stats:
        return []

    result = []
    for stat in stats:
        stat.setdefault("work_start_count", 0)
        stat.setdefault("work_end_count", 0)
        stat.setdefault("work_start_fines", 0)
        stat.setdefault("work_end_fines", 0)
        stat.setdefault("late_count", 0)
        stat.setdefault("early_count", 0)
        stat.setdefault("work_days", 0)
        stat.setdefault("work_hours", 0)

        if "activities" not in stat or not isinstance(stat["activities"], dict):
            stat["activities"] = {}

        result.append(stat)

    return result


async def get_monthly_stats_compatible(chat_id: int, target_date: date) -> List[Dict]:
    """兼容函数：获取月度统计数据并确保完整性"""
    try:
        month_start = target_date.replace(day=1)
        monthly_stats = await db.get_monthly_statistics(
            chat_id, month_start.year, month_start.month
        )

        if not monthly_stats:
            return []

        return await ensure_monthly_data_completeness(monthly_stats)

    except Exception as e:
        logger.error(f"获取兼容月度数据失败: {e}")
        return []


@admin_required
@rate_limit(rate=2, per=60)
async def cmd_exportmonthly(message: types.Message):
    """导出月度数据（使用新架构）"""
    args = message.text.split()
    chat_id = message.chat.id
    uid = message.from_user.id

    year = None
    month = None

    if len(args) >= 3:
        try:
            year = int(args[1])
            month = int(args[2])
            if month < 1 or month > 12:
                await message.answer(
                    "❌ 月份必须在1-12之间", reply_to_message_id=message.message_id
                )
                return
        except ValueError:
            await message.answer(
                "❌ 请输入有效的年份和月份", reply_to_message_id=message.message_id
            )
            return
    else:
        # 如果没有指定，默认导出上个月
        today = db.get_beijing_time()
        if today.month == 1:
            year = today.year - 1
            month = 12
        else:
            year = today.year
            month = today.month - 1

        await message.answer(
            f"📅 未指定月份，默认导出 {year}年{month}月 数据",
            reply_to_message_id=message.message_id,
        )

    await message.answer(
        f"⏳ 正在导出 {year}年{month}月 数据，请稍候...",
        reply_to_message_id=message.message_id,
    )

    try:
        from reset_service import _export_monthly_data_concurrent

        success = await _export_monthly_data_concurrent(
            chat_id=chat_id, year=year, month=month
        )

        if success:
            logger.info(
                f"👑 管理员 {uid} 成功导出群组 {chat_id} 的 {year}年{month}月 数据"
            )
            await message.answer(
                f"✅ {year}年{month}月 数据已导出并推送！",
                reply_to_message_id=message.message_id,
            )
        else:
            logger.error(
                f"❌ 管理员 {uid} 导出群组 {chat_id} 的 {year}年{month}月 数据失败"
            )
            await message.answer(
                f"❌ {year}年{month}月 数据导出失败，请检查日志",
                reply_to_message_id=message.message_id,
            )

    except Exception as e:
        logger.error(f"❌ 导出月度数据失败: {e}")
        logger.error(traceback.format_exc())
        await message.answer(
            f"❌ 导出月度数据失败：{e}", reply_to_message_id=message.message_id
        )


async def get_group_stats_from_monthly(chat_id: int, target_date: date) -> List[Dict]:
    """从月度统计表获取群组统计数据 - 优化版"""
    try:
        month_start = target_date.replace(day=1)

        logger.info(
            f"🔍 从月度表查询数据: 群组{chat_id}, 日期{target_date}, 月份{month_start}"
        )

        monthly_stats = await db.get_monthly_statistics(
            chat_id, month_start.year, month_start.month
        )

        if not monthly_stats:
            logger.warning(f"⚠️ 月度表中没有找到 {month_start} 的数据")
            return []

        # ---------- 获取所有活动名称（只执行一次） ----------
        activity_names = set()

        try:
            from database import db as database_db

            activity_limits = await database_db.get_activity_limits_cached()
            activity_names.update(activity_limits.keys())
        except Exception:
            activity_limits = {}

        result = []

        for stat in monthly_stats:
            raw_activities = stat.get("activities", {})

            # ---------- JSON解析 ----------
            if isinstance(raw_activities, str):
                try:
                    import json

                    raw_activities = json.loads(raw_activities)
                except Exception:
                    raw_activities = {}

            if isinstance(raw_activities, dict):
                activity_names.update(raw_activities.keys())

            formatted_activities = {}

            # ---------- 统一处理活动 ----------
            for act_name in activity_names:

                act_data = (
                    raw_activities.get(act_name, {})
                    if isinstance(raw_activities, dict)
                    else {}
                )

                count = 0
                time_val = 0

                if isinstance(act_data, dict):
                    count = act_data.get("count", act_data.get("activity_count", 0))
                    time_val = act_data.get("time", act_data.get("accumulated_time", 0))

                elif isinstance(act_data, (int, float)):
                    count = 1 if act_data > 0 else 0
                    time_val = act_data

                elif isinstance(act_data, str):
                    try:
                        import json

                        parsed = json.loads(act_data)
                        if isinstance(parsed, dict):
                            count = parsed.get("count", parsed.get("activity_count", 0))
                            time_val = parsed.get(
                                "time", parsed.get("accumulated_time", 0)
                            )
                    except Exception:
                        pass

                # ---------- 类型安全 ----------
                try:
                    count = int(float(count)) if count else 0
                except Exception:
                    count = 0

                try:
                    time_val = int(float(time_val)) if time_val else 0
                except Exception:
                    time_val = 0

                if count > 0 or time_val > 0:
                    formatted_activities[act_name] = {
                        "count": count,
                        "time": time_val,
                    }

            # ---------- fallback: 从字段恢复活动 ----------
            if not formatted_activities:

                exclude_fields = {
                    "user_id",
                    "nickname",
                    "shift",
                    "statistic_date",
                    "total_accumulated_time",
                    "total_activity_count",
                    "total_fines",
                    "overtime_count",
                    "total_overtime_time",
                    "work_days",
                    "work_hours",
                    "work_start_count",
                    "work_end_count",
                    "work_start_fines",
                    "work_end_fines",
                    "late_count",
                    "early_count",
                    "created_at",
                    "updated_at",
                    "id",
                    "chat_id",
                }

                for key, value in stat.items():

                    if (
                        key not in exclude_fields
                        and not key.endswith(("_count", "_time", "_fines"))
                        and isinstance(value, (int, float))
                        and value > 0
                    ):
                        formatted_activities[key] = {
                            "count": 0,
                            "time": int(value),
                        }

            user_data = {
                "user_id": stat.get("user_id", 0),
                "nickname": stat.get("nickname", f"用户{stat.get('user_id', 0)}"),
                "shift": stat.get("shift", "day"),
                "total_accumulated_time": stat.get("total_accumulated_time", 0),
                "total_activity_count": stat.get("total_activity_count", 0),
                "total_fines": stat.get("total_fines", 0),
                "overtime_count": stat.get("overtime_count", 0),
                "total_overtime_time": stat.get("total_overtime_time", 0),
                "work_days": stat.get("work_days", 0),
                "work_hours": stat.get("work_hours", 0),
                "work_start_count": stat.get("work_start_count", 0),
                "work_end_count": stat.get("work_end_count", 0),
                "work_start_fines": stat.get("work_start_fines", 0),
                "work_end_fines": stat.get("work_end_fines", 0),
                "late_count": stat.get("late_count", 0),
                "early_count": stat.get("early_count", 0),
                "activities": formatted_activities,
            }

            logger.debug(
                f"📊 用户 {user_data['user_id']} | "
                f"工作天数:{user_data['work_days']} "
                f"工作时长:{user_data['work_hours']}秒 "
                f"活动:{list(formatted_activities.keys())}"
            )

            result.append(user_data)

        logger.info(f"✅ 从月度表成功获取 {target_date} 数据，共 {len(result)} 个用户")

        return result

    except Exception as e:
        logger.error(f"❌ 从月度表获取数据失败: {e}")
        logger.error(traceback.format_exc())
        return []


async def export_and_push_csv(
    chat_id: int,
    to_admin_if_no_group: bool = True,
    file_name: str = None,
    target_date=None,
    is_daily_reset: bool = False,
    from_monthly_table: bool = False,
    push_file: bool = True,
) -> bool:
    """导出群组数据为 XLSX 并推送（整行空数据淡红色背景）"""

    # ===== 创建本地副本，避免作用域问题 =====
    local_chat_id = chat_id
    local_file_name = file_name
    local_target_date = target_date
    local_is_daily_reset = is_daily_reset
    local_from_monthly_table = from_monthly_table
    local_push_file = push_file
    local_to_admin_if_no_group = to_admin_if_no_group

    # ===== 创建看门狗 =====
    watchdog = Watchdog(timeout=300, name=f"export_{local_chat_id}")

    async def _export_impl():
        try:
            if not bot_manager or not bot_manager.bot:
                logger.error(f"❌ Bot管理器未初始化，无法导出 {local_chat_id}")
                if local_is_daily_reset:
                    return True
                return False

            if not await db._ensure_healthy_connection():
                logger.error(f"❌ 数据库连接不健康，无法导出 {local_chat_id}")
                if local_is_daily_reset:
                    return True
                return False

        except Exception as e:
            logger.error(f"❌ 前置检查失败 {local_chat_id}: {e}")
            if local_is_daily_reset:
                return True
            return False

        start_time = time.time()
        operation_id = f"export_{local_chat_id}_{int(start_time)}"
        logger.info(f"🚀 [{operation_id}] 开始导出群组 {local_chat_id} 的数据...")

        temp_file = None
        group_stats = []
        activity_limits = Config.DEFAULT_ACTIVITY_LIMITS.copy()

        try:
            await db.init_group(local_chat_id)

            # ========== 辅助函数定义 ==========
            def safe_int(value, default=0):
                """安全转换为整数"""
                if value is None:
                    return default
                try:
                    if isinstance(value, str):
                        if value.isdigit():
                            return int(value)
                        elif value.replace(".", "", 1).isdigit():
                            return int(float(value))
                        else:
                            return default
                    return int(value)
                except (ValueError, TypeError):
                    return default

            def safe_format_time(seconds):
                """安全格式化时间"""
                try:
                    return MessageFormatter.format_time_for_csv(safe_int(seconds))
                except Exception:
                    return "0分0秒"

            def format_shift_for_export(shift: str) -> str:
                """格式化班次为中文"""
                if not shift:
                    return "白班"
                shift_lower = str(shift).lower()
                if shift_lower == "day":
                    return "白班"
                if shift_lower in ["night", "night_last", "night_tonight"]:
                    return "夜班"
                return "白班"

            def format_export_value(value, is_time: bool = False):
                """格式化导出值，空数据显示为 '-'"""
                if value is None:
                    return "-"
                try:
                    num_value = int(value)
                    if num_value <= 0:
                        return "-"
                    if is_time:
                        return safe_format_time(num_value)
                    return str(num_value)
                except (ValueError, TypeError):
                    if not value or str(value).strip() == "":
                        return "-"
                    return str(value)

            def is_row_empty(row_data, exclude_indices=None):
                """判断整行数据是否为空（除了指定的排除列）"""
                if exclude_indices is None:
                    exclude_indices = {0, 1, 2}  # 排除用户ID、昵称、班次

                for idx, value in enumerate(row_data):
                    if idx in exclude_indices:
                        continue
                    if value != "-":
                        return False
                return True

            def create_excel_workbook(group_stats, all_activities, headers):
                """创建 Excel 工作簿并格式化"""
                from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
                from openpyxl.utils import get_column_letter

                # 创建工作簿
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "打卡统计"

                # 定义颜色
                HEADER_FILL = PatternFill(
                    start_color="4472C4", end_color="4472C4", fill_type="solid"
                )  # 蓝色标题
                HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
                DATA_FONT = Font(size=10)

                # ===== 新增：行背景色 =====
                HAS_FINE_FILL = PatternFill(
                    start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"
                )  # 淡黄色（有罚款）
                NO_FINE_FILL = PatternFill(
                    start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
                )  # 淡绿色（无罚款）
                EMPTY_ROW_FILL = PatternFill(
                    start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
                )  # 淡红色（完全空数据）

                BORDER = Border(
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                )

                # 写入表头
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border = BORDER

                ws.row_dimensions[1].height = 21

                # 先构建所有行数据
                all_rows_data = []
                for user_data in group_stats:
                    activities = user_data.get("activities", {})

                    # 获取罚款总额
                    total_fines = format_export_value(user_data.get("total_fines", 0))
                    # 判断是否有罚款（排除 "-" 和 0）
                    has_fine = total_fines != "-" and total_fines != "0"

                    # 构建行数据
                    row_data = [
                        user_data.get("user_id", "未知"),
                        user_data.get("nickname", "未知用户"),
                        format_shift_for_export(user_data.get("shift", "day")),
                        format_export_value(user_data.get("work_days", 0)),
                        format_export_value(user_data.get("work_start_count", 0)),
                        format_export_value(user_data.get("work_end_count", 0)),
                        format_export_value(
                            user_data.get("work_hours", 0), is_time=True
                        ),
                    ]

                    # 活动数据
                    for act in all_activities:
                        activity_info = activities.get(act, {})
                        row_data.append(
                            format_export_value(activity_info.get("count", 0))
                        )
                        row_data.append(
                            format_export_value(
                                activity_info.get("time", 0), is_time=True
                            )
                        )

                    # 统计数据
                    row_data.extend(
                        [
                            format_export_value(
                                user_data.get("total_activity_count", 0)
                            ),
                            format_export_value(
                                user_data.get("total_accumulated_time", 0), is_time=True
                            ),
                            format_export_value(user_data.get("overtime_count", 0)),
                            format_export_value(
                                user_data.get("total_overtime_time", 0), is_time=True
                            ),
                            format_export_value(user_data.get("early_count", 0)),
                            format_export_value(user_data.get("late_count", 0)),
                            format_export_value(user_data.get("work_end_fines", 0)),
                            format_export_value(user_data.get("work_start_fines", 0)),
                            total_fines,  # 罚款总金额
                        ]
                    )

                    all_rows_data.append({"data": row_data, "has_fine": has_fine})

                # 写入数据并应用样式
                for row_idx, row_info in enumerate(all_rows_data, 2):
                    row_data = row_info["data"]
                    has_fine = row_info["has_fine"]

                    # 判断整行是否为空（除了用户ID、昵称、班次）
                    is_empty = is_row_empty(row_data, exclude_indices={0, 1, 2})

                    # 确定行的背景色
                    if is_empty:
                        row_fill = EMPTY_ROW_FILL  # 淡红色 - 完全空数据
                    elif has_fine:
                        row_fill = HAS_FINE_FILL  # 淡黄色 - 有罚款
                    else:
                        row_fill = NO_FINE_FILL  # 淡绿色 - 无罚款

                    for col_idx, value in enumerate(row_data, 1):
                        cell = ws.cell(row=row_idx, column=col_idx, value=value)
                        cell.font = DATA_FONT
                        cell.border = BORDER
                        cell.alignment = Alignment(
                            horizontal="center", vertical="center"
                        )
                        cell.fill = row_fill

                    ws.row_dimensions[row_idx].height = 21

                # 自动调整列宽（优化版）
                for col_idx, col in enumerate(ws.columns, 1):
                    max_length = 0
                    column_letter = get_column_letter(col_idx)

                    for cell in col:
                        try:
                            if cell.value:
                                cell_value = str(cell.value)
                                cell_length = len(cell_value)

                                # 中文字符适当增加宽度
                                chinese_count = sum(
                                    1
                                    for char in cell_value
                                    if "\u4e00" <= char <= "\u9fff"
                                )
                                if chinese_count > 0:
                                    cell_length = cell_length + chinese_count * 1.2

                                max_length = max(max_length, cell_length)
                        except:
                            pass

                    adjusted_width = min(max(max_length + 2, 8), 120)
                    ws.column_dimensions[column_letter].width = adjusted_width

                # 冻结首行
                ws.freeze_panes = "A2"

                # 保存到 BytesIO
                output = BytesIO()
                wb.save(output)
                output.seek(0)

                return output

            # ========== 辅助函数定义结束 ==========

            beijing_now = db.get_beijing_time()
            current_hour = beijing_now.hour
            current_minute = beijing_now.minute
            current_time_decimal = current_hour + current_minute / 60

            shift_config = await db.get_shift_config(local_chat_id)
            day_start_str = shift_config.get("day_start", "09:00")
            day_start_hour = int(day_start_str.split(":")[0])
            day_start_minute = int(day_start_str.split(":")[1])
            day_start_decimal = day_start_hour + day_start_minute / 60

            watchdog.feed()

            # ===== 处理目标日期 =====
            working_target_date = local_target_date
            if working_target_date is not None:
                if hasattr(working_target_date, "date"):
                    working_target_date = working_target_date.date()
                elif not isinstance(working_target_date, date):
                    try:
                        if isinstance(working_target_date, str):
                            working_target_date = datetime.strptime(
                                working_target_date, "%Y-%m-%d"
                            ).date()
                    except Exception:
                        logger.warning(
                            f"⚠️ [{operation_id}] 无法解析target_date: {working_target_date}"
                        )
                        working_target_date = None

            if working_target_date is None:
                business_date = await db.get_business_date(local_chat_id)

                if local_is_daily_reset:
                    export_date = business_date - timedelta(days=1)
                    logger.info(
                        f"🔄 [{operation_id}] 自动重置导出: 业务日期={business_date}, "
                        f"导出日期={export_date}"
                    )
                else:
                    export_date = business_date
                    logger.info(
                        f"👤 [{operation_id}] 手动导出: 业务日期={business_date}, "
                        f"导出日期={export_date}"
                    )

                working_target_date = export_date
            else:
                logger.info(
                    f"📅 [{operation_id}] 使用指定的目标日期: {working_target_date}"
                )

            watchdog.feed()

            # ===== 生成文件名 =====
            current_file_name = local_file_name
            if not current_file_name:
                if local_is_daily_reset:
                    current_file_name = f"daily_backup_{local_chat_id}_{working_target_date:%Y%m%d}.xlsx"
                else:
                    current_file_name = f"manual_export_{local_chat_id}_{beijing_now:%Y%m%d_%H%M%S}.xlsx"

            logger.info(
                f"🔍 [{operation_id}] 获取群组 {local_chat_id} 的统计数据，日期: {working_target_date}"
            )

            # ===== 获取数据 =====
            current_from_monthly_table = local_from_monthly_table
            if current_from_monthly_table:
                logger.info(f"📊 [{operation_id}] 尝试从月度表获取数据")
                try:
                    group_stats = await get_group_stats_from_monthly(
                        local_chat_id, working_target_date
                    )
                    if group_stats:
                        logger.info(
                            f"✅ [{operation_id}] 从月度表获取到 {len(group_stats)} 条完整数据"
                        )
                        activity_limits = Config.DEFAULT_ACTIVITY_LIMITS.copy()
                    else:
                        logger.warning(f"⚠️ [{operation_id}] 月度表无数据，回退到常规表")
                        current_from_monthly_table = False
                except Exception as e:
                    logger.error(f"❌ [{operation_id}] 从月度表获取数据失败: {e}")
                    current_from_monthly_table = False

            watchdog.feed()

            if not current_from_monthly_table:
                try:
                    activity_task = asyncio.create_task(db.get_activity_limits_cached())
                    if local_is_daily_reset and working_target_date is not None:
                        stats_task = asyncio.create_task(
                            db.get_group_statistics_for_archive(
                                local_chat_id, working_target_date
                            )
                        )
                    else:
                        stats_task = asyncio.create_task(
                            db.get_group_statistics(
                                local_chat_id, working_target_date
                            )
                        )

                    results = await asyncio.gather(
                        activity_task, stats_task, return_exceptions=True
                    )

                    if isinstance(results[0], Exception):
                        logger.error(
                            f"❌ [{operation_id}] 获取活动配置失败: {results[0]}"
                        )
                        activity_limits = Config.DEFAULT_ACTIVITY_LIMITS.copy()
                    elif results[0]:
                        activity_limits = results[0]
                    else:
                        activity_limits = Config.DEFAULT_ACTIVITY_LIMITS.copy()

                    if isinstance(results[1], Exception):
                        logger.error(
                            f"❌ [{operation_id}] 获取统计数据失败: {results[1]}"
                        )
                        group_stats = []
                    elif results[1]:
                        group_stats = results[1]
                    else:
                        group_stats = []

                except Exception as e:
                    logger.error(f"❌ [{operation_id}] 并发获取数据失败: {e}")
                    activity_limits = Config.DEFAULT_ACTIVITY_LIMITS.copy()
                    group_stats = []

            watchdog.feed()

            if not group_stats:
                logger.warning(
                    f"⚠️ [{operation_id}] 群组 {local_chat_id} 没有数据需要导出"
                )
                if not local_is_daily_reset:
                    await bot_manager.send_message_with_retry(
                        local_chat_id, "⚠️ 当前没有数据需要导出"
                    )
                return True

            # ===== 数据验证和标准化 =====
            validated_stats = []
            for idx, user_data in enumerate(group_stats):
                if not isinstance(user_data, dict):
                    continue

                user_data["work_start_count"] = safe_int(
                    user_data.get("work_start_count", 0)
                )
                user_data["work_end_count"] = safe_int(
                    user_data.get("work_end_count", 0)
                )
                user_data["work_start_fines"] = safe_int(
                    user_data.get("work_start_fines", 0)
                )
                user_data["work_end_fines"] = safe_int(
                    user_data.get("work_end_fines", 0)
                )
                user_data["late_count"] = safe_int(user_data.get("late_count", 0))
                user_data["early_count"] = safe_int(user_data.get("early_count", 0))
                user_data["work_days"] = safe_int(user_data.get("work_days", 0))
                user_data["work_hours"] = safe_int(user_data.get("work_hours", 0))
                user_data["total_activity_count"] = safe_int(
                    user_data.get("total_activity_count", 0)
                )
                user_data["total_accumulated_time"] = safe_int(
                    user_data.get("total_accumulated_time", 0)
                )
                user_data["total_fines"] = safe_int(user_data.get("total_fines", 0))
                user_data["overtime_count"] = safe_int(
                    user_data.get("overtime_count", 0)
                )
                user_data["total_overtime_time"] = safe_int(
                    user_data.get("total_overtime_time", 0)
                )

                if "activities" not in user_data or not isinstance(
                    user_data["activities"], dict
                ):
                    user_data["activities"] = {}

                validated_stats.append(user_data)

                if idx % 10 == 0:
                    watchdog.feed()

            group_stats = validated_stats
            logger.info(
                f"📊 [{operation_id}] 数据验证完成，有效数据: {len(group_stats)} 条"
            )

            watchdog.feed()

            # ===== 获取所有活动并按字母排序 =====
            all_activities = sorted(activity_limits.keys())

            # ===== 定义表头 =====
            headers = [
                "用户ID",
                "用户昵称",
                "班次",
                "工作天数",
                "上班次数",
                "下班次数",
                "工作时长",
            ]

            for act in all_activities:
                headers.append(f"{act}次数")
                headers.append(f"{act}总时长")

            headers.extend(
                [
                    "活动次数总计",
                    "活动用时总计",
                    "超时次数",
                    "超时时长",
                    "早退次数",
                    "迟到次数",
                    "下班罚款",
                    "上班罚款",
                    "罚款总金额",
                ]
            )

            # ===== 统计信息 =====
            unique_users = set()
            total_records = 0
            for user_data in group_stats:
                user_id = user_data.get("user_id")
                if user_id:
                    unique_users.add(str(user_id))
                total_records += 1

            # ===== 创建 Excel 文件 =====
            excel_buffer = create_excel_workbook(
                group_stats=group_stats, all_activities=all_activities, headers=headers
            )

            # ===== 写入临时文件（异步写入 + fallback）=====
            temp_file = f"temp_{operation_id}_{current_file_name}"

            async def write_file_async():
                """异步写入文件，失败时降级为同步写入"""
                try:
                    async with aiofiles.open(temp_file, "wb") as f:
                        await f.write(excel_buffer.getvalue())
                    return True
                except Exception as e:
                    logger.error(f"❌ [{operation_id}] 异步写入文件失败: {e}")
                    try:
                        with open(temp_file, "wb") as f:
                            f.write(excel_buffer.getvalue())
                        return True
                    except Exception as sync_e:
                        logger.error(
                            f"❌ [{operation_id}] 同步写入文件也失败: {sync_e}"
                        )
                        return False

            async def get_chat_title_async():
                """异步获取群组标题"""
                try:
                    chat_info = await bot_manager.bot.get_chat(local_chat_id)
                    return chat_info.title or f"群组 {local_chat_id}"
                except Exception as e:
                    logger.warning(f"⚠️ [{operation_id}] 获取群组标题失败: {e}")
                    return f"群组 {local_chat_id}"

            # 并发执行：写入文件 + 获取群组标题
            write_result, chat_title = await asyncio.gather(
                write_file_async(), get_chat_title_async()
            )

            watchdog.feed()

            if not write_result:
                await bot_manager.send_message_with_retry(
                    local_chat_id, f"❌ 导出失败: 文件写入失败"
                )
                return False

            # ===== 发送文件 =====
            display_date = working_target_date.strftime("%Y年%m月%d日")
            dashed_line = getattr(
                MessageFormatter, "create_dashed_line", lambda: "─" * 30
            )()

            caption = (
                f"📊 <b>数据导出报告</b>\n"
                f"🏢 群组：<code>{chat_title}</code>\n"
                f"📅 统计日期：<code>{display_date}</code>\n"
                f"⏰ 导出时间：<code>{beijing_now.strftime('%Y-%m-%d %H:%M:%S')}</code>\n"
                f"{dashed_line}\n"
                f"💾 包含完整的工作记录统计（上班迟到/下班早退）\n"
                f"📈 总记录数: {total_records} 条\n"
                f"👥 总用户数: {len(unique_users)} 人\n"
                f"🎨 完全无活动记录的行已标注淡红色背景"
            )

            input_file = FSInputFile(temp_file, filename=current_file_name)
            send_to_group_success = False

            if local_push_file:
                try:
                    success = await bot_manager.send_document_with_retry(
                        chat_id=local_chat_id,
                        document=input_file,
                        caption=caption,
                        parse_mode="HTML",
                    )
                    if success:
                        send_to_group_success = True
                        logger.info(
                            f"✅ [{operation_id}] Excel文件已发送到群组 {local_chat_id}"
                        )
                    else:
                        logger.error(f"❌ [{operation_id}] bot_manager 发送文档失败")
                except Exception as e:
                    logger.error(f"❌ [{operation_id}] 发送到群组失败: {e}")
                    if not local_is_daily_reset:
                        await bot_manager.send_message_with_retry(
                            local_chat_id, f"❌ 数据导出失败: {str(e)[:100]}"
                        )
            else:
                logger.debug(f"⏭️ [{operation_id}] push_file=False，跳过文件发送")
                send_to_group_success = True

            if local_to_admin_if_no_group and notification_service:
                try:
                    await notification_service.send_document(
                        local_chat_id, input_file, caption=caption
                    )
                except Exception as e:
                    logger.warning(f"⚠️ [{operation_id}] 推送到通知服务失败: {e}")

            # 清理临时文件
            async def cleanup_background():
                await asyncio.sleep(2)
                if temp_file and os.path.exists(temp_file):
                    os.remove(temp_file)

            asyncio.create_task(cleanup_background())

            duration = time.time() - start_time
            logger.info(
                f"✅ [{operation_id}] 数据导出完成\n"
                f"   文件: {current_file_name}\n"
                f"   用户数: {len(unique_users)}, 数据行: {total_records}\n"
                f"   耗时: {duration:.2f}秒"
            )

            return send_to_group_success

        except Exception as e:
            logger.error(f"❌ [{operation_id}] 导出过程发生异常: {e}")
            logger.error(traceback.format_exc())

            try:
                await bot_manager.send_message_with_retry(
                    local_chat_id, f"❌ 数据导出失败: {str(e)[:100]}"
                )
            except:
                pass

            if temp_file and os.path.exists(temp_file):
                try:
                    os.remove(temp_file)
                except:
                    pass

            return False

    try:
        return await watchdog.run(_export_impl())
    except asyncio.CancelledError:
        logger.error(f"⏰ 导出操作超时，已取消 (chat_id={local_chat_id})")
        try:
            await bot_manager.send_message_with_retry(
                local_chat_id, "⏰ 导出操作超时，请重试"
            )
        except:
            pass
        return False

